import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side

# Filepath to file for material data and energy consumption, respectively
csvfilename = '//Users//Heimemappa//Desktop//Master//Fra Asanthi//prod_model_cell_material_data.csv'
csvfilename2 = '//Users//Heimemappa//Desktop//Master//Fra Asanthi//prod_model_energy.csv'


'''Creating the Excel file and sheets'''
wb = Workbook()
wb.active.title = 'Version'
wb.create_sheet('LOG')
wb.create_sheet('Foreground')
wb.create_sheet('F_f')
wb.create_sheet('A_bf')
wb.create_sheet('y_gen')
wb.create_sheet('Electricity_mixes')
wb.create_sheet('PRO')
wb.create_sheet('STR')


'''Version'''
ws = wb.active
ws['A1'] = 'Template Version:'
ws['B1'] = 1.1
ws['C1'] = '<----'
ws['D1'] = 'Do not edit!'


'''LOG'''
ws = wb['LOG']
ws['A2'] = 'Date:'
ws['B3'] = '14.07.2021'
ws['A5'] = 'Project:'
ws['B6'] = 'A coupled battery production and LCI model'
ws['A8'] = 'Authors:'
ws['B9'] = 'Dan André Johansen'
ws['E1'] = 'GENERAL INSTRUCTIONS'
ws['E3'] = '1) You can use this template in two situations:'
ws['F4'] = '1.1) You want to define a foreground, with a final demand (y_f), requirement matrices (A_ff, A_bf), ' \
           'and foreground emissions (F_f)'
ws['G5'] = "1.1.1) Fill in colored boxes in the sheets ''Foreground'', ''A_bf'' and ''F_f''"
ws['F6'] = '1.2) You only want to define a final demand vector, which calls for a final demand on generic data'
ws['G7'] = '1.2.1) Fill in colored boxes in the sheet y_gen'
ws['E12'] = '2) You can use white boxes for your own comments, these are not read by Arda'
ws['E13'] = '3) The process ID in the labels (Column C) MUST be numerical'
ws['E14'] = '4) The process ID int eh labels (Column C) MUST be unique, i.e. different from those of the background (' \
            'pick bigger numbers) '
ws['E15'] = '5) With only one exception, you are NOT allowed to modify the layout fo the template'
ws['F16'] = '5.1) Exception: You are allowed to add or remove columns between (but excluding) columns C and J of the ' \
            'sheet "Foreground" '
ws['F17'] = '5.2) In other words, you are allowed to modify the layout of the process labels "PRO_f" as long as:'
ws['G18'] = '5.2.1) …the first column of the label is the full NAME'
ws['G19'] = '5.2.2) …the second column of the label is the ID number'
ws['G20'] = '5.2.3) … the LAST column of the label is the unit'
ws['E21'] = "6) Do NOT edit the sheet named 'version'"
ws['E22'] = '7) Even though the coloring stops at some point, there should not be any limit to the size of any of the '\
            'matrices/vectors '
ws['E23'] = '8) Evidently, dimensions need to be consistent. '
ws['F24'] = 'In the sheet "Foreground": y_f and A_ff cannot have more rows than PRO_f'


'''Foreground'''
ws = wb['Foreground']
ws['A1'] = 'In this Sheet, you enter your foreground data: The process labels (green), the final demand vector (y_f, ' \
           'pink), and your foreground requirement matrix (A_ff, orange) '
ws['B2'] = 'Label (PRO_f):'
ws['K2'] = 'y_f:'
ws['M2'] = 'A_ff:'
ws['B3'] = 'FULL NAME'
ws['C3'] = 'PROCESS ID'
ws['D3'] = 'NAME'
ws['E3'] = 'Other ID'
ws['F3'] = 'INFRASTRUCTURE?'
ws['G3'] = 'LOCATION'
ws['H3'] = 'CATEGORY'
ws['I3'] = 'SUBCATEGORY'
ws['J3'] = 'UNIT'
ws['L4'] = 1  # the value of y_f

foreground_names = [
    'Battery cell',
    'Anode',
    'Cathode',
    'Separator',
    'Electrolyte',
    'Cell container',
    'Dry room',
    'Floor heating',
    'Formation',
    'Filling',
    'Stacking/Winding',
    'Slitting',
    'Calendering',
    'Drying',
    'Coating',
    'Mixing',
    'Negative current collector',
    'Positive current collector',
    'Negative electrode paste',
    'Positive electrode paste',
    'Anode tab',
    'Cathode tab',
    'Pouch',
    'Negative active material',
    'Positive active material',
    'Negative binder',
    'Positive binder',
    'LMO-G',
    'NMC811-G',
    'LFP G',
    'NCA G',
    'NMC622 G',
    'NMC111 G',
    'NMC532LMO',
    'NMC532 G',
    'LMO LTO',
    'NMC333 Si',
    'Precursor NCM',
    'Cobalt sulphate'
]

for column in range(14, len(foreground_names)+14):
    column_letter = get_column_letter(column)
    ws[column_letter+str(2)] = column-13
    ws[column_letter+str(3)] = foreground_names[column-14]

a = 4
b = a+len(foreground_names)
ws['C4'] = 1000001
for i in range(a, b):
    ws['A'+str(i)] = i-3
    ws['B'+str(i)] = foreground_names[i-a]
    ws['D'+str(i)] = foreground_names[i-a]
    if ws['C'+str(i)].value is None:
        ws['C'+str(i)] = ws['C4'].value+i+5

for i in range(5, 10):
    ws['N'+str(i)].value = 1

ws['O20'].value = 1
ws['P21'].value = 1

for i in range(24, 27):
    ws['S'+str(i)].value = 1

ws['AD22'].value = 1
ws['AE23'].value = 1
ws['AF27'].value = 1
ws['AF29'].value = 1
ws['AG28'].value = 1
ws['AG30'].value = 1

for i in range(4, 10):
    ws['J'+str(i)].value = 'pcs'
for i in range(10, 20):
    ws['J'+str(i)].value = 'kWh'
for i in range(20, ws.max_row+1):
    ws['J'+str(i)].value = 'pcs'

diag_fill = PatternFill(fill_type='solid', fgColor='FFE699')
other_fill = PatternFill(fill_type='solid', fgColor='FFC000')
y_ff_fill = PatternFill(fill_type='solid', fgColor='FF5B4C')
border_fill = Border(outline=Side(style='thick', color='808080'))

a = 4
b = ws.max_row+1
for i in range(a, b):
    for column in range(14, len(foreground_names) + 14):
        column_letter = get_column_letter(column)
        ws[column_letter + str(i)].fill = other_fill
        # ws2[letter+str(i)].border = border_fill

i = 4
for column in range(14, len(foreground_names) + 14):
    column_letter = get_column_letter(column)
    ws[column_letter+str(i)].fill = diag_fill
    i += 1

a = 4
b = ws.max_row+1
for i in range(a, b):
    ws['L'+str(i)].fill = y_ff_fill

ws['K2'].fill = y_ff_fill
ws['M2'].fill = other_fill


'''F_f part 1'''
ws = wb['F_f']

# Arda IDs
battery_cell_ardaIDs = [
    509,
    964,
    177,
    387
]

precursor_NCM_ardaIDs = 720

cobalt_sulphate_ardaIDs = [
    599,
    800,
    26386,
    118,
    240,
    396,
    400,
    404,
    1168,
    1187,
    1241,
    1247,
    1282,
    1288,
    1235,
    1297,
    1301,
    1373,
    1383,
    1389,
    1399,
    1428,
    24675,
    1543,
    1533,
    1604
]

ws['A1'] = 'In this sheet, you enter direct stressor emissions of the foreground. The indexes will be assembled as an '\
           'F_f (also called S_f) matrix. '
ws['A3'] = 'STRESSOR NAME'
ws['B3'] = 'FOREGROUND PROCESS NAME'
ws['C3'] = '(Matrix row)'
ws['D3'] = '(Matrix column)'
ws['E3'] = '(Value)'
ws['F3'] = 'UNIT'
ws['A4'] = 'Comment'
ws['B4'] = 'Comment'
ws['C4'] = 'STRESSOR ROW #'
ws['D4'] = 'FOREGROUND PROCESS ID #'
ws['E4'] = 'AMOUNT'
ws['F4'] = 'Comment'
ws['G4'] = 'Comment'
ws['H4'] = 'Comment'

'''
for i, bat in enumerate(battery_cell_ardaIDs):
    ws2['A'+str(i+5)] = '=VLOOKUP(C'+str(i+5)+',STR!$B$2:$E$25951,4,FALSE)'
    ws2['B'+str(i+5)] = '=Foreground!$B$3'
    ws2['C'+str(i+5)] = bat
    ws2['D'+str(i+5)] = '=Foreground!$C$3'
    ws2['F'+str(i+5)] = '=VLOOKUP(C'+str(i+5)+',STR!$B$2:$M$25951,12,FALSE)'

# a = ws2.max_row+1
# ws2['A'+str(a)] = '=VLOOKUP(C'+str(a)+',STR!$B$2:$E$25951,4,FALSE)'
# ws2['B9'] = 'Foreground!$B$'+ ????    HAR IKKE PRECURSOR NCM MED I FORELØPIG LCI-VERSJON
# ws2['C9'] = precursor_NCM_ardaIDs     HAR IKKE PRECURSOR NCM MED I FORELØPIG LCI-VERSJON

# a = ws2.max_row+1
# for i, cobalt in enumerate(cobalt_sulphate_ardaIDs):
'''


'''Background (A_bf)'''
ws = wb['A_bf']

# Variables and Arda IDs
electricity_sources = [
    'Coal',
    'Oil',
    'Natural Gas',
    'Nuclear',
    'Hydropower',
    'Solar PV',
    'Wind',
    'Waste Heat',
    'Biofuels',
    'Other Sources'
]

values_from_Ellingsen = [
    0.0000000188,       # [0] Precious metal refinery - Battery cell
    0.0000000004,       # [1] Chemical factory - Electrolyte, (-) paste, (+) paste, (+) active material
    50 * 27000000,      # [2] Plastic processing factory - Separator, Pouch
    50 * 130000000,     # [3] Aluminium casting facility - (+) CC, Cathode tab, Pouch
    0.000000000458,     # [4] Metal working factory - (-) CC, Anode tab
    0.2,                # [5] Transport: Freight train - Several components
    0.1,                # [6] Transport: Lorry >32 ton - Several components
    0.6,                # [7] Transport: Freight train - Several components
    0.020005635,        # [8] Share of CMC and PAA to battery grade graphite - negative electrode paste
    1.25,               # [9] Sodium hydroxide - NCA
    0.27,               # [10] Lithium hydroxide - NCA
    0.49,               # [11] Nickel - NCA
    0.04,               # [12] Aluminium hydroxide - NCA
    0.09,               # [13] Cobalt sulphate - NCA
    0.65,               # [14] Phosphoric acid - LFP
    0.46                # [15] Lithium hydroxide - LFP
]

values_from_MajeauBettez = [
    0.550130568209837,      # [0] Heat in chemical industry
    0.718572637228255,      # [1] Transport: Freight train
    0.119762106204709,      # [2] Transport: Lorry>32 ton
    0.0000000004,           # [3] Chemical factory
    0.876802678869194,      # [4] Soda ash - Precursor NCM
    1.53608259379624,       # [5] Transport: Freight train - Precursor NCM
    0.256013765632706,      # [6] Transport: Lorry >32 ton - Precursor NCM
    0.0323190821944853,     # [7] Chemicals, inorganic - Cobalt sulphate
    0.00950561241014273,    # [8] Chemicals, organic - Cobalt sulphate
    0.00149428227087444,    # [9] Hydrogen cyanide - Cobalt sulphate
    0.0190112248202855,     # [10] Limestone - Cobalt sulphate
    1.37641267698867,       # [11] Cement - Cobalt sulphate
    17.3382370361003,       # [12] Sand - Cobalt sulphate
    0.0631172664033477,     # [13] Blasting - Cobalt sulphate
    4.60071640650908,       # [14] Diesel - Cobalt sulphate
    1.78325288814278,       # [15] Electricity - Cobalt sulphate
    3.42202046765138E-10,   # [16] Aluminium hydroxide - Cobalt sulphate
    1.59694288490398E-06,   # [17] Conveyor belt - Cobalt sulphate
    2.0912347302314E-09,    # [18] Mine infrastructure - Cobalt sulphate
    0.942956751086158,      # [19] Transport: Lorry >32 ton - Cobalt sulphate
    13.2698349245592,       # [20] Non-sulfidic overburden - Cobalt sulphate
    24.7145922663711        # [21] Non-sulfidic tailings - Cobalt sulphate
]

el_mix_and_transmission_network = [
    9138,   # Coal
    11120,  # Oil
    9527,   # Natural Gas
    9628,   # Nuclear
    9292,   # Hydropower
    11265,  # Solar
    9748,   # Wind
    11954,  # Waste Heat
    12026,  # Biofuels
    3666,   # Transmission network
    3668,   # Transmission network
    3667,   # Transmission network
    600     # Sulfur
]

bat_cell = [3724]

sulfur = 600

transport = [
    13687,  # Transport: Freight train
    2809    # Transport: Lorry >32 ton
]

electrolyte = [
    701,    # Ethylene carbonate (EC). 'DMC or EMC' is also added here as I couldn't find explicit values for those.
    6501,   # LiPF6
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3641    # Chemical factory
]

cathode = [
    13687,  # Transport: Freight train
    2809    # Transport: Lorry >32 ton
]

anode = [
    13687,  # Transport: Freight train
    2809    # Transport: Lorry >32 ton
]

cell_container = [
    13687,  # Transport: Freight train
    2809    # Transport: Lorry >32 ton
]

separator = [
    7036,   # Plastic PE
    2683,   # Injection moulding
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    4006    # Plastic processing factory
]

positive_cc_Al = [
    7880,   # Aluminium
    1953,   # Sheet rolling, aluminium
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3713    # Aluminium casting facility
]

positive_electrode_paste = [
    6317,   # Solvent - NMP
    465,    # Carbon black
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3641    # Chemical factory
]

negative_cc_Cu = [
    1797,   # Copper
    12219,  # Scrap copper
    1955,   # Sheet rolling, copper
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3741    # Metal working factory
]

negative_electrode_paste = [
    2948,   # Carboxymethyl cellulose (CMC)
    649,    # Acrylic acid
    3389,   # Solvent - Water
    6317,   # Solvent - NMP
    465,    # Carbon black
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3641    # Chemical factory
]

tab_Al = [
    7880,   # Aluminium
    1953,   # Sheet rolling, aluminium
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3713    # Aluminium casting facility
]

tab_Cu = [
    1797,   # Copper
    12219,  # Scrap copper
    1955,   # Sheet rolling, copper
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3741    # Metal working factory
]

pouch_Al = [
    7958,   # Aluminium
    2653,   # Plastic PET
    2647,   # Nylon 6
    2662,   # Plastic PP
    2684,   # Packaging film
    2683,   # Injection moulding
    1953,   # Sheet rolling, aluminium
    3713,   # Aluminium casting facility
    4006,   # Plastic processing plant
    13687,  # Transport: Freight train
    2809    # Transport: Lorry >32 ton
]

# Positive active material
positive_active_material = [
    521,    # Lithium hydroxide
    8656,   # LMO
    728,    # Heat in chemical industry
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3641    # Chemical factory
]

positive_active_chemistry = [
    8656,   # LMO
    521,    # NMC811
    2,      # LFP
    3,      # NCA
    521,    # NMC622
    521,    # NMC111
    521,    # NMC532LMO
    521,    # NMC532
    8,      # LMO LTO
    521     # NMC333 Si
]

precursor_NCM = [
    6293,   # Manganese sulfate
    6530,   # Nickel sulfate
    567,    # Soda, powder
    13687,  # Transport: Freight train
    2809,   # Transport: Lorry >32 ton
    3641    # Chemical plant
]

cobalt_sulphate = [
    469,    # Chemicals inorganic
    673,    # Chemicals organic
    733,    # Hydrogen cyanide
    880,    # Limestone
    7606,   # Cement
    830,    # Sand
    892,    # Blasting
    896,    # Diesel, burned in building machine
    10817,  # Electricity, medium voltage
    3715,   # Aluminium hydroxide
    3660,   # Conveyor belt
    3738,   # Non-ferrous metal mine
    2809,   # Transport: Lorry >32 ton
    3301,   # Disposal, non-sulfidic overburden
    3302    # Disposal, non-sulfidic tailings
]

binders = [
    777,  # PVDF
    6382  # CMC
]

negative_active_material = [
    491,  # Battery grade graphite
    1749  # Silicon nano-wire
]

transmission_values = [
    0.00000000844,
    0.000000000317,
    0.0000000324
]

transmission_values_sulfur = 0.000000075425

NCA = [
    6612,   # Sodium hydroxide
    521,    # Lithium hydroxide
    1857,   # Nickel (Was listed as 521 in Ellingsen, but I corrected it here)
    6338,   # Aluminium hydroxide
    728,    # Heat in chemical industry
    13695,  # Transport: Freight train
    13830,  # Transport: Lorry >32 ton
    3641    # Chemical plant
]

LFP = [
    1831,   # Iron sulfate
    539,    # Phosphoric acid
    521,    # Lithium hydroxide
    13695,  # Transport: Freight train
    13830,  # Transport: Lorry >32 ton
    314    # Chemical factory
]

# Battery cell
a = 5
b = a + len(bat_cell)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$4'
    ws['C' + str(i)] = bat_cell[i - 5]
    ws['D' + str(i)] = '=Foreground!$C$4'

# Separator
a = ws.max_row + 1
b = a + len(separator)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$7'
    ws['C' + str(i)] = separator[i - a]
    ws['D' + str(i)] = '=Foreground!$C$7'

# Electrolyte
a = ws.max_row + 1
b = a + len(electrolyte)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$8'
    ws['C' + str(i)] = electrolyte[i - a]
    ws['D' + str(i)] = '=Foreground!$C$8'

# Dry room
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 10
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=Foreground!$N$' + str(c) + '*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Floor heating
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 11
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=Foreground!$N$' + str(c) + '*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Formation
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 12
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=Foreground!$N$' + str(c) + '*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Filling
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 13
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=Foreground!$N$' + str(c) + '*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Stacking/Winding
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 14
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=Foreground!$N$' + str(c) + '*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Slitting
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 15
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=(Foreground!$O$' + str(c) + '+Foreground!$P$' + str(c) + ')*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Calendering
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 16
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=(Foreground!$O$' + str(c) + '+Foreground!$P$' + str(c) + ')*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Drying
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 17
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=(Foreground!$O$' + str(c) + '+Foreground!$P$' + str(c) + ')*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Coating
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 18
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=(Foreground!$O$' + str(c) + '+Foreground!$P$' + str(c) + ')*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Mixing
j = 7
a = ws.max_row + 1
b = a + len(el_mix_and_transmission_network)
c = 19
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$' + str(c)
    ws['C' + str(i)] = el_mix_and_transmission_network[i - a]
    ws['D' + str(i)] = '=Foreground!$C$' + str(c)
    if i <= a + 8:
        ws['E' + str(i)] = '=(Foreground!$O$' + str(c) + '+Foreground!$P$' + str(c) + ')*J' + str(j) + '*$M$7*$M$8'
        j += 1
    if i == b - 1:
        ws['E' + str(i)] = '=' + str(transmission_values_sulfur) + '*(J' + str(a + 9) + ')'
        ws['I' + str(a + 8)] = 'HV'
        ws['J' + str(a + 8)] = '=SUM(E' + str(a) + ':E' + str(a + 8) + ')'
        ws['I' + str(a + 9)] = 'MV'
        ws['J' + str(a + 9)] = '=J' + str(a + 8) + '/(M7*M8)'
        ws['B' + str(b - 1)] = '=Foreground!$B$' + str(c)
        for j, trans in enumerate(transmission_values):
            ws['E' + str(j + a + 9)] = '=' + str(trans) + '*J' + str(a + 8)

# Negative current collector Cu
a = ws.max_row + 1
b = a + len(negative_cc_Cu)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$20'
    ws['C' + str(i)] = negative_cc_Cu[i - a]
    ws['D' + str(i)] = '=Foreground!$C$20'

# Positive current collector Al
a = ws.max_row + 1
b = a + len(positive_cc_Al)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$21'
    ws['C' + str(i)] = positive_cc_Al[i - a]
    ws['D' + str(i)] = '=Foreground!$C$21'

# Negative electrode paste
a = ws.max_row + 1
b = a + len(negative_electrode_paste)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$22'
    ws['C' + str(i)] = negative_electrode_paste[i - a]
    ws['D' + str(i)] = '=Foreground!$C$22'

# Positive electrode paste
a = ws.max_row + 1
b = a + len(positive_electrode_paste)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$23'
    ws['C' + str(i)] = positive_electrode_paste[i - a]
    ws['D' + str(i)] = '=Foreground!$C$23'

# Anode tab
a = ws.max_row + 1
b = a + len(tab_Cu)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$24'
    ws['C' + str(i)] = tab_Cu[i - a]
    ws['D' + str(i)] = '=Foreground!$C$24'

# Cathode tab
a = ws.max_row + 1
b = a + len(tab_Al)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$25'
    ws['C' + str(i)] = tab_Al[i - a]
    ws['D' + str(i)] = '=Foreground!$C$25'

# Pouch
a = ws.max_row + 1
b = a + len(pouch_Al)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$26'
    ws['C' + str(i)] = pouch_Al[i - a]
    ws['D' + str(i)] = '=Foreground!$C$26'

# Negative active material
a = ws.max_row + 1
b = a + len(negative_active_material)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$27'
    ws['C' + str(i)] = negative_active_material[i - a]
    ws['D' + str(i)] = '=Foreground!$C$27'

# Negative binder
a = ws.max_row + 1
b = a + len(binders)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$29'
    ws['C' + str(i)] = binders[i - a]
    ws['D' + str(i)] = '=Foreground!$C$29'

# Positive binder
a = ws.max_row + 1
b = a + len(binders)
for i in range(a, b):
    ws['B' + str(i)] = '=Foreground!$B$30'
    ws['C' + str(i)] = binders[i - a]
    ws['D' + str(i)] = '=Foreground!$C$30'

# Import of data from process model
wb2 = Workbook()
ws2 = wb2.active

with open(csvfilename) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    for row in csv_reader:
        ws2.append(row)

j = 22
a = -25
for i in range(2, ws2.max_row + 1):
    if ws2['A' + str(i)].value == ws2['A' + str(i - 1)].value:
        ws2['F' + str(a)].value = ws2['A' + str(i)].value
        if ws2['B' + str(i)].value == 'base':
            ws2['G' + str(a)].value = ws2['B' + str(i)].value
            ws2['F' + str(j)].value = ws2['C' + str(i)].value
            if ws2['G' + str(j)].value != 'base':
                ws2['G' + str(j)].value = float(ws2['D' + str(i)].value)
            j += 1
        elif ws2['B' + str(i)].value == 'base-scrap':
            ws2['H' + str(a)].value = ws2['B' + str(i)].value
            if ws2['G' + str(j - 23)].value != 'base':
                ws2['H' + str(j - 23)].value = float(ws2['D' + str(i)].value)
            ws2['I' + str(a)].value = 'SUM (total input)'
            # ws2['I'+str(j-23)].value = '=G'+str(j-23)+'+H'+str(j-23)
            ws2['I' + str(j - 23)].value = ws2['G' + str(j - 23)].value + ws2['H' + str(j - 23)].value
            j += 1
    else:
        a += 26
        j += -20

c = 0
f = -24
g = -24
d = 0
anode3 = 0
cathode3 = 0
separator3 = 0
electrolyte3 = 0
container3 = 0
for i in range(2, ws2.max_row + 1):
    if ws2['B' + str(i)].value == 'cell':
        d = c + float(ws2['D' + str(i)].value)
        c = d
        f += 2
        if 'anode' in ws2['C' + str(i)].value:
            anode2 = anode3 + float(ws2['D' + str(i)].value)
            anode3 = anode2
            g += 2
        elif 'cathode' in ws2['C' + str(i)].value:
            cathode2 = cathode3 + float(ws2['D' + str(i)].value)
            cathode3 = cathode2
            g += 2
        elif 'separator' in ws2['C' + str(i)].value:
            separator2 = separator3 + float(ws2['D' + str(i)].value)
            separator3 = separator2
            g += 2
        elif 'electrolyte' in ws2['C' + str(i)].value:
            electrolyte2 = electrolyte3 + float(ws2['D' + str(i)].value)
            electrolyte3 = electrolyte2
            g += 2
        elif 'container' or 'terminal' in ws2['C' + str(i)].value:
            container2 = container3 + float(ws2['D' + str(i)].value)
            container3 = container2
            g += 2
    elif ws2['B' + str(i)].value == 'base':
        ws2['K' + str(f - 1)].value = 'Cell weight'
        ws2['K' + str(f)].value = d
        c = 0
        ws2['M' + str(g - 1)].value = 'Component'
        ws2['M' + str(g)].value = 'Anode'
        ws2['M' + str(g + 1)].value = 'Cathode'
        ws2['M' + str(g + 2)].value = 'Separator'
        ws2['M' + str(g + 3)].value = 'Electrolyte'
        ws2['M' + str(g + 4)].value = 'Container'
        ws2['N' + str(g - 1)].value = 'Mass'
        ws2['N' + str(g)].value = anode3
        ws2['N' + str(g + 1)].value = cathode3
        ws2['N' + str(g + 2)].value = separator3
        ws2['N' + str(g + 3)].value = electrolyte3
        ws2['N' + str(g + 4)].value = container3
    else:
        anode3 = 0
        cathode3 = 0
        separator3 = 0
        electrolyte3 = 0
        container3 = 0

wb3 = Workbook()
ws3 = wb3.active

with open(csvfilename2) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    for row in csv_reader:
        ws3.append(row)

f = 2
e = 0
for i in range(2, ws3.max_row + 1):
    ws3['F' + str(i)].value = float(ws3['C' + str(i)].value)
    if e != 15:
        ws3['G' + str(i)].value = ws3['F' + str(i)].value * ws2['K' + str(f)].value / 3.6
        e += 1
    else:
        f += 26
        ws3['G' + str(i)].value = ws3['F' + str(i)].value * ws2['K' + str(f)].value / 3.6
        e = 1

# Choice of battery chemistry
battery_chemistry = input('NMC811\nNMC622\nNMC111\nNMC532\nNMC532LMO\nNMC333SI\nLFP\nLMO\nNCA\nSelect one of the '
                          'chemistries listed above: ')
if battery_chemistry.upper() not in ('NMC811', 'NMC622', 'NMC111', 'NMC532', 'NMC532LMO', 'NMC333SI', 'LFP', 'LMO',
                                     'NCA'):
    print('WARNING: The selected chemistry is not among the available choices. Please check for typos. '
          'Program stopped.')
    exit()

if battery_chemistry.upper() == 'LMO':

    # Energy consumption
    c = list(range(2, 17))
    wb['Foreground']['N10'] = ws3['G'+str(c[14])].value
    wb['Foreground']['N11'] = ws3['G'+str(c[13])].value
    wb['Foreground']['N12'] = ws3['G'+str(c[12])].value
    wb['Foreground']['N13'] = ws3['G'+str(c[11])].value
    wb['Foreground']['N14'] = ws3['G'+str(c[10])].value
    wb['Foreground']['P15'] = ws3['G'+str(c[9])].value
    wb['Foreground']['O15'] = ws3['G'+str(c[8])].value
    wb['Foreground']['P16'] = ws3['G'+str(c[7])].value
    wb['Foreground']['O16'] = ws3['G'+str(c[6])].value
    wb['Foreground']['P17'] = ws3['G'+str(c[5])].value
    wb['Foreground']['O17'] = ws3['G'+str(c[4])].value
    wb['Foreground']['P18'] = ws3['G'+str(c[3])].value
    wb['Foreground']['O18'] = ws3['G'+str(c[2])].value
    wb['Foreground']['P19'] = ws3['G'+str(c[1])].value
    wb['Foreground']['O19'] = ws3['G'+str(c[0])].value

    # positive active materials
    a = ws.max_row + 1
    b = a + len(positive_active_material)
    for i in range(a, b):
        ws['B' + str(i)] = '=Foreground!$B$31'
        ws['C' + str(i)] = positive_active_material[i - a]
        ws['D' + str(i)] = '=Foreground!$C$31'

    a = list(range(2, 25))
    wb['Foreground']['AL31'] = 1

    # Mass of active material
    ws['R6'] = ws2['I' + str(a[0])].value

    # Masses of main components
    ws['P7'].value = ws2['N' + str(a[0])].value
    ws['P8'].value = ws2['N' + str(a[1])].value
    ws['P9'].value = ws2['N' + str(a[2])].value
    ws['P10'].value = ws2['N' + str(a[3])].value
    ws['P11'].value = ws2['N' + str(a[4])].value

    for i in range(5, ws.max_row + 1):

        # Separator
        if ws['B' + str(i)].value == '=Foreground!$B$7':

            # Plastic PE
            if ws['C' + str(i)].value == separator[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value
            # Injection moulding
            elif ws['C' + str(i)].value == separator[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == separator[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == separator[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value * values_from_Ellingsen[6]
            # Plastic processing factory
            elif ws['C' + str(i)].value == separator[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value / values_from_Ellingsen[2]

        # Electrolyte
        elif ws['B' + str(i)].value == '=Foreground!$B$8':

            # Ethylene carbonate (EC)
            if ws['C' + str(i)].value == electrolyte[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[15])].value + ws2['I' + str(a[16])].value
            # LiPF6
            elif ws['C' + str(i)].value == electrolyte[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[14])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == electrolyte[2]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == electrolyte[3]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[6]
            # Chemical factory
            elif ws['C' + str(i)].value == electrolyte[4]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[1]

        # Copper - Foil (Negative current collector)
        elif ws['B' + str(i)].value == '=Foreground!$B$20':

            # Copper
            if ws['C' + str(i)].value == negative_cc_Cu[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[9])].value
            # Scrap copper
            elif ws['C' + str(i)].value == negative_cc_Cu[1]:
                ws['E' + str(i)].value = 0
            # Sheet rolling, copper
            elif ws['C' + str(i)].value == negative_cc_Cu[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[9])].value + 0
            # Transport: Freight train
            elif ws['C' + str(i)].value == negative_cc_Cu[3]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == negative_cc_Cu[4]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[6]
            # Metal working factory
            elif ws['C' + str(i)].value == negative_cc_Cu[5]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[4]

        # Aluminum - Foil (Positive current collector)
        elif ws['B' + str(i)].value == '=Foreground!$B$21':

            # Aluminium
            if ws['C' + str(i)].value == positive_cc_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == positive_cc_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value
            # # Transport: Freight train
            elif ws['C' + str(i)].value == positive_cc_Al[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_cc_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value * values_from_Ellingsen[6]
            # Aluminium casting facility
            elif ws['C' + str(i)].value == positive_cc_Al[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value / values_from_Ellingsen[3]

        # Negative electrode paste
        elif ws['B' + str(i)].value == '=Foreground!$B$22':

            # Carboxymethyl cellulose
            if ws['C' + str(i)].value == negative_electrode_paste[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value * values_from_Ellingsen[8]
            # Acrylic acid (PAA)
            elif ws['C' + str(i)].value == negative_electrode_paste[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value * values_from_Ellingsen[8]
            # Water - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[22])].value
            # NMP - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[20])].value
            # Carbon black - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[3])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == negative_electrode_paste[5]:
                ws['E' + str(i)].value = (2 * ws2['I' + str(a[1])].value * values_from_Ellingsen[8] +
                                          ws2['I' + str(a[1])].value + ws2['I' + str(a[22])].value) * \
                                         values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == negative_electrode_paste[6]:
                ws['E' + str(i)].value = (2 * ws2['I' + str(a[1])].value * values_from_Ellingsen[8] +
                                          ws2['I' + str(a[1])].value + ws2['I' + str(a[22])].value) * \
                                         values_from_Ellingsen[6]

        # Positive electrode paste
        elif ws['B' + str(i)].value == '=Foreground!$B$23':

            # NMP - Cathode
            if ws['C' + str(i)].value == positive_electrode_paste[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[21])].value
            # Carbon black - Cathode
            elif ws['C' + str(i)].value == positive_electrode_paste[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[4])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == positive_electrode_paste[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[1]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_electrode_paste[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[2]

        # Anode tab
        elif ws['B' + str(i)].value == '=Foreground!$B$24':

            # Copper
            if ws['C' + str(i)].value == tab_Cu[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value
            # Scrap copper
            elif ws['C' + str(i)].value == tab_Cu[1]:
                ws['E' + str(i)].value = 0
            # Sheet rolling, copper
            elif ws['C' + str(i)].value == tab_Cu[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == tab_Cu[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == tab_Cu[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[6]
            # Metal working factory
            elif ws['C' + str(i)].value == tab_Cu[5]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[4]

        # Cathode tab
        elif ws['B' + str(i)].value == '=Foreground!$B$25':

            # Aluminium
            if ws['C' + str(i)].value == tab_Al[0]:
                ws['E' + str(i)].value = ws2['G' + str(a[12])].value + ws2['H' + str(a[12])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == tab_Al[1]:
                ws['E' + str(i)].value = ws2['G' + str(a[12])].value + ws2['H' + str(a[12])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == tab_Al[2]:
                ws['E' + str(i)].value = (ws2['G' + str(a[12])].value + ws2['H' + str(a[12])].value) * \
                                         values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == tab_Al[3]:
                ws['E' + str(i)].value = (ws2['G' + str(a[12])].value + ws2['H' + str(a[12])].value) * \
                                         values_from_Ellingsen[6]
            # Aluminium casting facility
            elif ws['C' + str(i)].value == tab_Al[4]:
                ws['E' + str(i)].value = (ws2['G' + str(a[12])].value + ws2['H' + str(a[12])].value) / \
                                         values_from_Ellingsen[3]

        # Container
        elif ws['B' + str(i)].value == '=Foreground!$B$26':

            # Aluminum - Pouch
            if ws['C' + str(i)].value == pouch_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value
            # Plastic PET
            elif ws['C' + str(i)].value == pouch_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[19])].value
            # Nylon 6
            elif ws['C' + str(i)].value == pouch_Al[2]:
                ws['E' + str(i)].value = 0
            # Plastic PP
            elif ws['C' + str(i)].value == pouch_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[18])].value
            # Packaging film
            elif ws['C' + str(i)].value == pouch_Al[4]:
                ws['E' + str(i)].value = 0
            # Injection moulding
            elif ws['C' + str(i)].value == pouch_Al[5]:
                ws['E' + str(i)].value = ws2['I' + str(a[19])].value + ws2['I' + str(a[18])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == pouch_Al[6]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value
            # Aluminium casting facility
            elif ws['C' + str(i)].value == pouch_Al[7]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value / values_from_Ellingsen[3]
            # Plastic processing factory
            elif ws['C' + str(i)].value == pouch_Al[8]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value) / \
                                         values_from_Ellingsen[2]
            # Transport: Freight train
            elif ws['C' + str(i)].value == pouch_Al[9]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value +
                                          ws2['I' + str(a[13])].value) * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == pouch_Al[10]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value +
                                          ws2['I' + str(a[13])].value) * values_from_Ellingsen[6]

        # Negative active material
        elif ws['B' + str(i)].value == '=Foreground!$B$27':

            # Battery grade graphite - Anode
            if ws['C' + str(i)].value == negative_active_material[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value
            # Silicon nano-wire - Anode
            elif ws['C' + str(i)].value == negative_active_material[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[2])].value

        # Negative binder
        elif ws['B' + str(i)].value == '=Foreground!$B$29':

            # PVDF
            if ws['C' + str(i)].value == binders[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[7])].value
            # CMC
            elif ws['C' + str(i)].value == binders[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[5])].value

        # Positive binder
        elif ws['B' + str(i)].value == '=Foreground!$B$30':
            # PVDF
            if ws['C' + str(i)].value == binders[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[8])].value
            # CMC
            elif ws['C' + str(i)].value == binders[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[6])].value

        # Positive active material
        elif ws['D' + str(i)].value == '=Foreground!$C$31':

            # Lithium hydroxide
            if ws['C' + str(i)].value == positive_active_material[0]:
                ws['E' + str(i)].value = 0
            # LMO
            if ws['C' + str(i)].value == positive_active_material[1]:
                ws['E' + str(i)].value = ws['R6'].value
            # Heat in chemical industry
            if ws['C' + str(i)].value == positive_active_material[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[0]
            # Transport: Freight train
            if ws['C' + str(i)].value == positive_active_material[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[1]
            # Transport: Lorry >32 ton
            if ws['C' + str(i)].value == positive_active_material[4]:
                ws['E' + str(i)].value = values_from_MajeauBettez[2]
            # Chemical factory
            if ws['C' + str(i)].value == positive_active_material[5]:
                ws['E' + str(i)].value = values_from_Ellingsen[1]

if battery_chemistry.upper() in ('NMC811', 'NMC622', 'NMC111', 'NMC532LMO', 'NMC532', 'NMC333SI'):

    # positive active materials. The c-variable is used for the energy consumption below
    a = ws.max_row + 1
    b = a + len(positive_active_material)
    for i in range(a, b):
        ws['C' + str(i)] = positive_active_material[i - a]
        if battery_chemistry.upper() == 'NMC811':
            ws['B' + str(i)] = '=Foreground!$B$32'
            ws['D' + str(i)] = '=Foreground!$C$32'
            c = list(range(17, 32))
        elif battery_chemistry.upper() == 'NMC622':
            ws['B' + str(i)] = '=Foreground!$B$35'
            ws['D' + str(i)] = '=Foreground!$C$35'
            c = list(range(62, 77))
        elif battery_chemistry.upper() == 'NMC111':
            ws['B' + str(i)] = '=Foreground!$B$36'
            ws['D' + str(i)] = '=Foreground!$C$36'
            c = list(range(77, 92))
        elif battery_chemistry.upper() == 'NMC532LMO':
            ws['B' + str(i)] = '=Foreground!$B$37'
            ws['D' + str(i)] = '=Foreground!$C$37'
            c = list(range(92, 107))
        elif battery_chemistry.upper() == 'NMC532':
            ws['B' + str(i)] = '=Foreground!$B$38'
            ws['D' + str(i)] = '=Foreground!$C$38'
            c = list(range(107, 122))
        elif battery_chemistry.upper() == 'NMC333SI':
            ws['B' + str(i)] = '=Foreground!$B$40'
            ws['D' + str(i)] = '=Foreground!$C$40'
            c = list(range(137, 152))

    # Energy consumption
    wb['Foreground']['N10'] = ws3['G'+str(c[14])].value
    wb['Foreground']['N11'] = ws3['G'+str(c[13])].value
    wb['Foreground']['N12'] = ws3['G'+str(c[12])].value
    wb['Foreground']['N13'] = ws3['G'+str(c[11])].value
    wb['Foreground']['N14'] = ws3['G'+str(c[10])].value
    wb['Foreground']['P15'] = ws3['G'+str(c[9])].value
    wb['Foreground']['O15'] = ws3['G'+str(c[8])].value
    wb['Foreground']['P16'] = ws3['G'+str(c[7])].value
    wb['Foreground']['O16'] = ws3['G'+str(c[6])].value
    wb['Foreground']['P17'] = ws3['G'+str(c[5])].value
    wb['Foreground']['O17'] = ws3['G'+str(c[4])].value
    wb['Foreground']['P18'] = ws3['G'+str(c[3])].value
    wb['Foreground']['O18'] = ws3['G'+str(c[2])].value
    wb['Foreground']['P19'] = ws3['G'+str(c[1])].value
    wb['Foreground']['O19'] = ws3['G'+str(c[0])].value
    wb['Foreground']['AY42'] = 1

    # precursor NCM
    a = ws.max_row + 1
    b = a + len(precursor_NCM)
    for i in range(a, b):
        ws['B' + str(i)] = '=Foreground!$B$41'
        ws['C' + str(i)] = precursor_NCM[i - a]
        ws['D' + str(i)] = '=Foreground!$C$41'

    # cobalt sulphate
    a = ws.max_row + 1
    b = a + len(cobalt_sulphate)
    for i in range(a, b):
        ws['B' + str(i)] = '=Foreground!$B$42'
        ws['C' + str(i)] = cobalt_sulphate[i - a]
        ws['D' + str(i)] = '=Foreground!$C$42'

    # Molecular masses (source: https://www.webqc.org/molecular-weight-of-Li.html)
    M_Li = 6.9410
    M_O2 = 31.999
    M_Ni = 58.693
    M_Mn = 54.938
    M_Co = 58.933

    if battery_chemistry.upper() == 'NMC811':
        a = list(range(28, 51))
        b = list(range(17, 32))
        wb['Foreground']['AL32'] = 1
        wb['Foreground']['AP41'] = 1

        # Mass of active material
        ws['R6'] = ws2['G' + str(a[0])].value + ws2['H' + str(a[0])].value

        # Molar fractions
        y_Ni = 0.8 / 4
        y_Mn = 0.1 / 4
        y_Co = 0.1 / 4

    elif battery_chemistry.upper() == 'NMC622':
        a = list(range(106, 129))
        wb['Foreground']['AL35'] = 1
        wb['Foreground']['AS41'] = 1

        # Mass of active material
        ws['R6'] = ws2['I' + str(a[0])].value

        # Molar fractions
        y_Ni = 0.6 / 4
        y_Mn = 0.2 / 4
        y_Co = 0.2 / 4

    elif battery_chemistry.upper() == 'NMC111':
        a = list(range(132, 155))
        wb['Foreground']['AL36'] = 1
        wb['Foreground']['AT41'] = 1

        # Mass of active material
        ws['R6'] = ws2['I' + str(a[0])].value

        # Molar fractions
        y_Ni = 1 / 3 / 4
        y_Mn = 1 / 3 / 4
        y_Co = 1 / 3 / 4

    elif battery_chemistry.upper() == 'NMC532LMO':
        a = list(range(158, 181))
        wb['Foreground']['AL37'] = 1
        wb['Foreground']['AU41'] = 1

        # Mass of active material
        ws['R6'] = ws2['I' + str(a[0])].value

        # Molar fractions
        y_Ni = 0.5 / 4
        y_Mn = 0.3 / 4
        y_Co = 0.2 / 4

    elif battery_chemistry.upper() == 'NMC532':
        a = list(range(184, 207))
        wb['Foreground']['AL38'] = 1
        wb['Foreground']['AV41'] = 1

        # Mass of active material
        ws['R6'] = ws2['I' + str(a[0])].value

        # Molar fractions
        y_Ni = 0.5 / 4
        y_Mn = 0.3 / 4
        y_Co = 0.2 / 4

    elif battery_chemistry.upper() == 'NMC333SI':
        a = list(range(236, 259))
        wb['Foreground']['AL40'] = 1
        wb['Foreground']['AX41'] = 1

        # Mass of active material
        ws['R6'] = ws2['I' + str(a[0])].value

        # Molar fractions
        y_Ni = 1 / 3 / 4
        y_Mn = 1 / 3 / 4
        y_Co = 1 / 3 / 4

    # Mass fractions (the statement below is to avoid Python from warning about a potential error)
    # noinspection PyUnboundLocalVariable
    m_tot = M_Li + M_O2 + M_Ni * y_Ni + M_Mn * y_Mn + M_Co * y_Co
    mf_LiO2 = (M_Li + M_O2) / m_tot
    mf_Ni = M_Ni * y_Ni / m_tot
    mf_Mn = M_Mn * y_Mn / m_tot
    mf_Co = M_Co * y_Co / m_tot

    # Masses of main components
    ws['P7'].value = ws2['N' + str(a[0])].value
    ws['P8'].value = ws2['N' + str(a[1])].value
    ws['P9'].value = ws2['N' + str(a[2])].value
    ws['P10'].value = ws2['N' + str(a[3])].value
    ws['P11'].value = ws2['N' + str(a[4])].value

    for i in range(5, ws.max_row + 1):

        # Separator
        if ws['B' + str(i)].value == '=Foreground!$B$7':

            # Plastic PE
            if ws['C' + str(i)].value == separator[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value
            # Injection moulding
            elif ws['C' + str(i)].value == separator[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == separator[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == separator[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value * values_from_Ellingsen[6]
            # Plastic processing factory
            elif ws['C' + str(i)].value == separator[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value / values_from_Ellingsen[2]

        # Electrolyte
        elif ws['B' + str(i)].value == '=Foreground!$B$8':

            # Ethylene carbonate (EC)
            if ws['C' + str(i)].value == electrolyte[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[15])].value + ws2['I' + str(a[16])].value
            # LiPF6
            elif ws['C' + str(i)].value == electrolyte[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[14])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == electrolyte[2]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == electrolyte[3]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[6]
            # Chemical factory
            elif ws['C' + str(i)].value == electrolyte[4]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[1]

        # Copper - Foil (Negative current collector)
        elif ws['B' + str(i)].value == '=Foreground!$B$20':

            # Copper
            if ws['C' + str(i)].value == negative_cc_Cu[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[9])].value
            # Scrap copper
            elif ws['C' + str(i)].value == negative_cc_Cu[1]:
                ws['E' + str(i)].value = 0
            # Sheet rolling, copper
            elif ws['C' + str(i)].value == negative_cc_Cu[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[9])].value + 0
            # Transport: Freight train
            elif ws['C' + str(i)].value == negative_cc_Cu[3]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == negative_cc_Cu[4]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[6]
            # Metal working factory
            elif ws['C' + str(i)].value == negative_cc_Cu[5]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[4]

        # Aluminum - Foil (Positive current collector)
        elif ws['B' + str(i)].value == '=Foreground!$B$21':

            # Aluminium
            if ws['C' + str(i)].value == positive_cc_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == positive_cc_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value
            # # Transport: Freight train
            elif ws['C' + str(i)].value == positive_cc_Al[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_cc_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value * values_from_Ellingsen[6]
            # Aluminium casting facility
            elif ws['C' + str(i)].value == positive_cc_Al[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value / values_from_Ellingsen[3]

        # Negative electrode paste
        elif ws['B' + str(i)].value == '=Foreground!$B$22':

            # Carboxymethyl cellulose
            if ws['C' + str(i)].value == negative_electrode_paste[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value * values_from_Ellingsen[8]
            # Acrylic acid (PAA)
            elif ws['C' + str(i)].value == negative_electrode_paste[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value * values_from_Ellingsen[8]
            # Water - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[22])].value
            # NMP - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[20])].value
            # Carbon black - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[3])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == negative_electrode_paste[5]:
                ws['E' + str(i)].value = (2 * ws2['I' + str(a[1])].value * values_from_Ellingsen[8] +
                                          ws2['I' + str(a[1])].value + ws2['I' + str(a[22])].value) * \
                                         values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == negative_electrode_paste[6]:
                ws['E' + str(i)].value = (2 * ws2['I' + str(a[1])].value * values_from_Ellingsen[8] +
                                          ws2['I' + str(a[1])].value + ws2['I' + str(a[22])].value) * \
                                         values_from_Ellingsen[6]

        # Positive electrode paste
        elif ws['B' + str(i)].value == '=Foreground!$B$23':

            # NMP - Cathode
            if ws['C' + str(i)].value == positive_electrode_paste[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[21])].value
            # Carbon black - Cathode
            elif ws['C' + str(i)].value == positive_electrode_paste[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[4])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == positive_electrode_paste[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[1]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_electrode_paste[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[2]

        # Anode tab
        elif ws['B' + str(i)].value == '=Foreground!$B$24':

            # Copper
            if ws['C' + str(i)].value == tab_Cu[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value
            # Scrap copper
            elif ws['C' + str(i)].value == tab_Cu[1]:
                ws['E' + str(i)].value = 0
            # Sheet rolling, copper
            elif ws['C' + str(i)].value == tab_Cu[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == tab_Cu[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == tab_Cu[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[6]
            # Metal working factory
            elif ws['C' + str(i)].value == tab_Cu[5]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[4]

        # Cathode tab
        elif ws['B' + str(i)].value == '=Foreground!$B$25':

            # Aluminium
            if ws['C' + str(i)].value == tab_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == tab_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == tab_Al[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == tab_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value * values_from_Ellingsen[6]
            # Aluminium casting facility
            elif ws['C' + str(i)].value == tab_Al[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value / values_from_Ellingsen[3]

        # Container
        elif ws['B' + str(i)].value == '=Foreground!$B$26':

            # Aluminum - Pouch
            if ws['C' + str(i)].value == pouch_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value
            # Plastic PET
            elif ws['C' + str(i)].value == pouch_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[19])].value
            # Nylon 6
            elif ws['C' + str(i)].value == pouch_Al[2]:
                ws['E' + str(i)].value = 0
            # Plastic PP
            elif ws['C' + str(i)].value == pouch_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[18])].value
            # Packaging film
            elif ws['C' + str(i)].value == pouch_Al[4]:
                ws['E' + str(i)].value = 0
            # Injection moulding
            elif ws['C' + str(i)].value == pouch_Al[5]:
                ws['E' + str(i)].value = ws2['I' + str(a[19])].value + ws2['I' + str(a[18])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == pouch_Al[6]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value
            # Aluminium casting facility
            elif ws['C' + str(i)].value == pouch_Al[7]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value / values_from_Ellingsen[3]
            # Plastic processing factory
            elif ws['C' + str(i)].value == pouch_Al[8]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value) / \
                                         values_from_Ellingsen[2]
            # Transport: Freight train
            elif ws['C' + str(i)].value == pouch_Al[9]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value +
                                          ws2['I' + str(a[13])].value) * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == pouch_Al[10]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value +
                                          ws2['I' + str(a[13])].value) * values_from_Ellingsen[6]

        # Negative active material
        elif ws['B' + str(i)].value == '=Foreground!$B$27':

            # Battery grade graphite - Anode
            if ws['C' + str(i)].value == negative_active_material[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value
            # Silicon nano-wire - Anode
            elif ws['C' + str(i)].value == negative_active_material[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[2])].value

        # Negative binder
        elif ws['B' + str(i)].value == '=Foreground!$B$29':

            # PVDF
            if ws['C' + str(i)].value == binders[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[7])].value
            # CMC
            elif ws['C' + str(i)].value == binders[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[5])].value

        # Positive binder
        elif ws['B' + str(i)].value == '=Foreground!$B$30':
            # PVDF
            if ws['C' + str(i)].value == binders[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[8])].value
            # CMC
            elif ws['C' + str(i)].value == binders[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[6])].value

        # Positive active material
        if ws['D' + str(i)].value == '=Foreground!$C$37':

            mf_LiO2 = mf_LiO2 * 0.5
            mf_Ni = mf_Ni * 0.5
            mf_Mn = mf_Mn * 0.5
            mf_Co = mf_Co * 0.5
            mf_LMO = 0.5

            # Lithium hydroxide
            if ws['C' + str(i)].value == positive_active_material[0]:
                ws['E' + str(i)].value = mf_LiO2 * ws['R6'].value
            # LMO
            elif ws['C' + str(i)].value == positive_active_material[1]:
                ws['E' + str(i)].value = mf_LMO * ws['R6'].value
            # Heat in chemical industry
            elif ws['C' + str(i)].value == positive_active_material[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[0]
            # Transport: Freight train
            elif ws['C' + str(i)].value == positive_active_material[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[1]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_active_material[4]:
                ws['E' + str(i)].value = values_from_MajeauBettez[2]
            # Chemical factory
            elif ws['C' + str(i)].value == positive_active_material[5]:
                ws['E' + str(i)].value = values_from_Ellingsen[1]

        elif ws['D' + str(i)].value in ('=Foreground!$C$32', '=Foreground!$C$35', '=Foreground!$C$36',
                                        '=Foreground!$C$38', '=Foreground!$C$40'):

            # Lithium hydroxide
            if ws['C' + str(i)].value == positive_active_material[0]:
                ws['E' + str(i)].value = mf_LiO2 * ws['R6'].value
            # LMO
            elif ws['C' + str(i)].value == positive_active_material[1]:
                ws['E' + str(i)].value = 0
            # Heat in chemical industry
            elif ws['C' + str(i)].value == positive_active_material[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[0]
            # Transport: Freight train
            elif ws['C' + str(i)].value == positive_active_material[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[1]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_active_material[4]:
                ws['E' + str(i)].value = values_from_MajeauBettez[2]
            # Chemical factory
            elif ws['C' + str(i)].value == positive_active_material[5]:
                ws['E' + str(i)].value = values_from_Ellingsen[1]

        # Precursor NCM
        if ws['D' + str(i)].value == '=Foreground!$C$41':

            # Manganese sulfate
            if ws['C' + str(i)].value == precursor_NCM[0]:
                ws['E' + str(i)].value = mf_Mn * ws['R6'].value
            # Nickel sulfate
            if ws['C' + str(i)].value == precursor_NCM[1]:
                ws['E' + str(i)].value = mf_Co * ws['R6'].value
            # Soda ash
            if ws['C' + str(i)].value == precursor_NCM[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[4]
            # Transport: Freight train
            if ws['C' + str(i)].value == precursor_NCM[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[5]
            # Transport: Lorry >32 ton
            if ws['C' + str(i)].value == precursor_NCM[4]:
                ws['E' + str(i)].value = values_from_MajeauBettez[6]
            # Chemical factory
            if ws['C' + str(i)].value == precursor_NCM[5]:
                ws['E' + str(i)].value = values_from_MajeauBettez[3]

        # Cobalt sulphate
        if ws['D' + str(i)].value == '=Foreground!$C$42':

            # Chemicals, inorganic
            if ws['C' + str(i)].value == cobalt_sulphate[0]:
                ws['E' + str(i)].value = values_from_MajeauBettez[7]
            # Chemicals, organic
            elif ws['C' + str(i)].value == cobalt_sulphate[1]:
                ws['E' + str(i)].value = values_from_MajeauBettez[8]
            # Hydrogen cyanide
            elif ws['C' + str(i)].value == cobalt_sulphate[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[9]
            # Lime
            elif ws['C' + str(i)].value == cobalt_sulphate[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[10]
            # Cement
            elif ws['C' + str(i)].value == cobalt_sulphate[4]:
                ws['E' + str(i)].value = values_from_MajeauBettez[11]
            # Sand/Gravel
            elif ws['C' + str(i)].value == cobalt_sulphate[5]:
                ws['E' + str(i)].value = values_from_MajeauBettez[12]
            # Blasting
            elif ws['C' + str(i)].value == cobalt_sulphate[6]:
                ws['E' + str(i)].value = values_from_MajeauBettez[13]
            # Diesel
            elif ws['C' + str(i)].value == cobalt_sulphate[7]:
                ws['E' + str(i)].value = values_from_MajeauBettez[14]
            # Electricity
            elif ws['C' + str(i)].value == cobalt_sulphate[8]:
                ws['E' + str(i)].value = values_from_MajeauBettez[15]
            # Aluminium hydroxide
            elif ws['C' + str(i)].value == cobalt_sulphate[9]:
                ws['E' + str(i)].value = values_from_MajeauBettez[16]
            # Conveyor belt
            elif ws['C' + str(i)].value == cobalt_sulphate[10]:
                ws['E' + str(i)].value = values_from_MajeauBettez[17]
            # Mine infrastructure
            elif ws['C' + str(i)].value == cobalt_sulphate[11]:
                ws['E' + str(i)].value = values_from_MajeauBettez[18]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == cobalt_sulphate[12]:
                ws['E' + str(i)].value = values_from_MajeauBettez[19]
            # Non-sulfidic overburden
            elif ws['C' + str(i)].value == cobalt_sulphate[13]:
                ws['E' + str(i)].value = values_from_MajeauBettez[20]
            # Non-sulfidic tailing
            elif ws['C' + str(i)].value == cobalt_sulphate[14]:
                ws['E' + str(i)].value = values_from_MajeauBettez[21]

if battery_chemistry.upper() == 'LFP':

    # Positive active materials
    a = ws.max_row + 1
    b = a + len(LFP)
    for i in range(a, b):
        ws['B' + str(i)] = '=Foreground!$B$33'
        ws['C' + str(i)] = LFP[i - a]
        ws['D' + str(i)] = '=Foreground!$C$33'

    # Energy consumption
    c = list(range(32, 47))
    wb['Foreground']['N10'] = ws3['G'+str(c[14])].value
    wb['Foreground']['N11'] = ws3['G'+str(c[13])].value
    wb['Foreground']['N12'] = ws3['G'+str(c[12])].value
    wb['Foreground']['N13'] = ws3['G'+str(c[11])].value
    wb['Foreground']['N14'] = ws3['G'+str(c[10])].value
    wb['Foreground']['P15'] = ws3['G'+str(c[9])].value
    wb['Foreground']['O15'] = ws3['G'+str(c[8])].value
    wb['Foreground']['P16'] = ws3['G'+str(c[7])].value
    wb['Foreground']['O16'] = ws3['G'+str(c[6])].value
    wb['Foreground']['P17'] = ws3['G'+str(c[5])].value
    wb['Foreground']['O17'] = ws3['G'+str(c[4])].value
    wb['Foreground']['P18'] = ws3['G'+str(c[3])].value
    wb['Foreground']['O18'] = ws3['G'+str(c[2])].value
    wb['Foreground']['P19'] = ws3['G'+str(c[1])].value
    wb['Foreground']['O19'] = ws3['G'+str(c[0])].value

    a = list(range(54, 77))
    wb['Foreground']['AL33'] = 1

    # Mass of active material
    ws['R6'] = ws2['I' + str(a[0])].value

    # Masses of main components
    ws['P7'].value = ws2['N' + str(a[0])].value
    ws['P8'].value = ws2['N' + str(a[1])].value
    ws['P9'].value = ws2['N' + str(a[2])].value
    ws['P10'].value = ws2['N' + str(a[3])].value
    ws['P11'].value = ws2['N' + str(a[4])].value

    for i in range(5, ws.max_row + 1):

        # Separator
        if ws['B' + str(i)].value == '=Foreground!$B$7':

            # Plastic PE
            if ws['C' + str(i)].value == separator[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value
            # Injection moulding
            elif ws['C' + str(i)].value == separator[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == separator[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == separator[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value * values_from_Ellingsen[6]
            # Plastic processing factory
            elif ws['C' + str(i)].value == separator[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value / values_from_Ellingsen[2]

        # Electrolyte
        elif ws['B' + str(i)].value == '=Foreground!$B$8':

            # Ethylene carbonate (EC)
            if ws['C' + str(i)].value == electrolyte[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[15])].value + ws2['I' + str(a[16])].value
            # LiPF6
            elif ws['C' + str(i)].value == electrolyte[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[14])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == electrolyte[2]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == electrolyte[3]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[6]
            # Chemical factory
            elif ws['C' + str(i)].value == electrolyte[4]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[1]

        # Copper - Foil (Negative current collector)
        elif ws['B' + str(i)].value == '=Foreground!$B$20':

            # Copper
            if ws['C' + str(i)].value == negative_cc_Cu[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[9])].value
            # Scrap copper
            elif ws['C' + str(i)].value == negative_cc_Cu[1]:
                ws['E' + str(i)].value = 0
            # Sheet rolling, copper
            elif ws['C' + str(i)].value == negative_cc_Cu[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[9])].value + 0
            # Transport: Freight train
            elif ws['C' + str(i)].value == negative_cc_Cu[3]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == negative_cc_Cu[4]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[6]
            # Metal working factory
            elif ws['C' + str(i)].value == negative_cc_Cu[5]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[4]

        # Aluminum - Foil (Positive current collector)
        elif ws['B' + str(i)].value == '=Foreground!$B$21':

            # Aluminium
            if ws['C' + str(i)].value == positive_cc_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == positive_cc_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value
            # # Transport: Freight train
            elif ws['C' + str(i)].value == positive_cc_Al[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_cc_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value * values_from_Ellingsen[6]
            # Aluminium casting facility
            elif ws['C' + str(i)].value == positive_cc_Al[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value / values_from_Ellingsen[3]

        # Negative electrode paste
        elif ws['B' + str(i)].value == '=Foreground!$B$22':

            # Carboxymethyl cellulose
            if ws['C' + str(i)].value == negative_electrode_paste[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value * values_from_Ellingsen[8]
            # Acrylic acid (PAA)
            elif ws['C' + str(i)].value == negative_electrode_paste[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value * values_from_Ellingsen[8]
            # Water - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[22])].value
            # NMP - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[20])].value
            # Carbon black - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[3])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == negative_electrode_paste[5]:
                ws['E' + str(i)].value = (2 * ws2['I' + str(a[1])].value * values_from_Ellingsen[8] +
                                          ws2['I' + str(a[1])].value + ws2['I' + str(a[22])].value) * \
                                         values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == negative_electrode_paste[6]:
                ws['E' + str(i)].value = (2 * ws2['I' + str(a[1])].value * values_from_Ellingsen[8] +
                                          ws2['I' + str(a[1])].value + ws2['I' + str(a[22])].value) * \
                                         values_from_Ellingsen[6]

        # Positive electrode paste
        elif ws['B' + str(i)].value == '=Foreground!$B$23':

            # NMP - Cathode
            if ws['C' + str(i)].value == positive_electrode_paste[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[21])].value
            # Carbon black - Cathode
            elif ws['C' + str(i)].value == positive_electrode_paste[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[4])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == positive_electrode_paste[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[1]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_electrode_paste[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[2]

        # Anode tab
        elif ws['B' + str(i)].value == '=Foreground!$B$24':

            # Copper
            if ws['C' + str(i)].value == tab_Cu[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value
            # Scrap copper
            elif ws['C' + str(i)].value == tab_Cu[1]:
                ws['E' + str(i)].value = 0
            # Sheet rolling, copper
            elif ws['C' + str(i)].value == tab_Cu[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == tab_Cu[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == tab_Cu[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[6]
            # Metal working factory
            elif ws['C' + str(i)].value == tab_Cu[5]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[4]

        # Cathode tab
        elif ws['B' + str(i)].value == '=Foreground!$B$25':

            # Aluminium
            if ws['C' + str(i)].value == tab_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == tab_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == tab_Al[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == tab_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value * values_from_Ellingsen[6]
            # Aluminium casting facility
            elif ws['C' + str(i)].value == tab_Al[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value / values_from_Ellingsen[3]

        # Container
        elif ws['B' + str(i)].value == '=Foreground!$B$26':

            # Aluminum - Pouch
            if ws['C' + str(i)].value == pouch_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value
            # Plastic PET
            elif ws['C' + str(i)].value == pouch_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[19])].value
            # Nylon 6
            elif ws['C' + str(i)].value == pouch_Al[2]:
                ws['E' + str(i)].value = 0
            # Plastic PP
            elif ws['C' + str(i)].value == pouch_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[18])].value
            # Packaging film
            elif ws['C' + str(i)].value == pouch_Al[4]:
                ws['E' + str(i)].value = 0
            # Injection moulding
            elif ws['C' + str(i)].value == pouch_Al[5]:
                ws['E' + str(i)].value = ws2['I' + str(a[19])].value + ws2['I' + str(a[18])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == pouch_Al[6]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value
            # Aluminium casting facility
            elif ws['C' + str(i)].value == pouch_Al[7]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value / values_from_Ellingsen[3]
            # Plastic processing factory
            elif ws['C' + str(i)].value == pouch_Al[8]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value) / \
                                         values_from_Ellingsen[2]
            # Transport: Freight train
            elif ws['C' + str(i)].value == pouch_Al[9]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value +
                                          ws2['I' + str(a[13])].value) * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == pouch_Al[10]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value +
                                          ws2['I' + str(a[13])].value) * values_from_Ellingsen[6]

        # Negative active material
        elif ws['B' + str(i)].value == '=Foreground!$B$27':

            # Battery grade graphite - Anode
            if ws['C' + str(i)].value == negative_active_material[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value
            # Silicon nano-wire - Anode
            elif ws['C' + str(i)].value == negative_active_material[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[2])].value

        # Negative binder
        elif ws['B' + str(i)].value == '=Foreground!$B$29':

            # PVDF
            if ws['C' + str(i)].value == binders[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[7])].value
            # CMC
            elif ws['C' + str(i)].value == binders[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[5])].value

        # Positive binder
        elif ws['B' + str(i)].value == '=Foreground!$B$30':
            # PVDF
            if ws['C' + str(i)].value == binders[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[8])].value
            # CMC
            elif ws['C' + str(i)].value == binders[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[6])].value

        # Positive active material
        elif ws['D' + str(i)].value == '=Foreground!$C$33':

            # Iron sulfate
            if ws['C' + str(i)].value == LFP[0]:
                ws['E' + str(i)].value = ws['R6'].value
            # Phosphoric acid
            elif ws['C' + str(i)].value == LFP[1]:
                ws['E' + str(i)].value = ws['R6'].value * values_from_Ellingsen[14]
            # Lithium hydroxide
            elif ws['C' + str(i)].value == LFP[2]:
                ws['E' + str(i)].value = ws['R6'].value * values_from_Ellingsen[15]
            # Transport: Freight train
            elif ws['C' + str(i)].value == LFP[3]:
                ws['E' + str(i)].value = ws['R6'].value * (1+values_from_Ellingsen[14]+values_from_Ellingsen[15]) \
                                         * values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == LFP[4]:
                ws['E' + str(i)].value = ws['R6'].value * (1+values_from_Ellingsen[14]+values_from_Ellingsen[15]) \
                                         * values_from_Ellingsen[6]
            # Chemical factory
            elif ws['C' + str(i)].value == LFP[5]:
                ws['E' + str(i)].value = values_from_Ellingsen[1]

if battery_chemistry.upper() == 'NCA':

    # Positive active materials
    a = ws.max_row + 1
    b = a + len(NCA)
    for i in range(a, b):
        ws['B' + str(i)] = '=Foreground!$B$34'
        ws['C' + str(i)] = NCA[i - a]
        ws['D' + str(i)] = '=Foreground!$C$34'

    # cobalt sulphate
    a = ws.max_row + 1
    b = a + len(cobalt_sulphate)
    for i in range(a, b):
        ws['B' + str(i)] = '=Foreground!$B$42'
        ws['C' + str(i)] = cobalt_sulphate[i - a]
        ws['D' + str(i)] = '=Foreground!$C$42'

    # Energy consumption
    c = list(range(47, 62))
    wb['Foreground']['N10'] = ws3['G'+str(c[14])].value
    wb['Foreground']['N11'] = ws3['G'+str(c[13])].value
    wb['Foreground']['N12'] = ws3['G'+str(c[12])].value
    wb['Foreground']['N13'] = ws3['G'+str(c[11])].value
    wb['Foreground']['N14'] = ws3['G'+str(c[10])].value
    wb['Foreground']['P15'] = ws3['G'+str(c[9])].value
    wb['Foreground']['O15'] = ws3['G'+str(c[8])].value
    wb['Foreground']['P16'] = ws3['G'+str(c[7])].value
    wb['Foreground']['O16'] = ws3['G'+str(c[6])].value
    wb['Foreground']['P17'] = ws3['G'+str(c[5])].value
    wb['Foreground']['O17'] = ws3['G'+str(c[4])].value
    wb['Foreground']['P18'] = ws3['G'+str(c[3])].value
    wb['Foreground']['O18'] = ws3['G'+str(c[2])].value
    wb['Foreground']['P19'] = ws3['G'+str(c[1])].value
    wb['Foreground']['O19'] = ws3['G'+str(c[0])].value

    a = list(range(80, 103))
    wb['Foreground']['AL34'] = 1
    wb['Foreground']['AR42'] = 1

    # Mass of active material
    ws['R6'] = ws2['I' + str(a[0])].value

    # Masses of main components
    ws['P7'].value = ws2['N' + str(a[0])].value
    ws['P8'].value = ws2['N' + str(a[1])].value
    ws['P9'].value = ws2['N' + str(a[2])].value
    ws['P10'].value = ws2['N' + str(a[3])].value
    ws['P11'].value = ws2['N' + str(a[4])].value

    for i in range(5, ws.max_row + 1):

        # Separator
        if ws['B' + str(i)].value == '=Foreground!$B$7':

            # Plastic PE
            if ws['C' + str(i)].value == separator[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value
            # Injection moulding
            elif ws['C' + str(i)].value == separator[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == separator[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == separator[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value * values_from_Ellingsen[6]
            # Plastic processing factory
            elif ws['C' + str(i)].value == separator[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[17])].value / values_from_Ellingsen[2]

        # Electrolyte
        elif ws['B' + str(i)].value == '=Foreground!$B$8':

            # Ethylene carbonate (EC)
            if ws['C' + str(i)].value == electrolyte[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[15])].value + ws2['I' + str(a[16])].value
            # LiPF6
            elif ws['C' + str(i)].value == electrolyte[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[14])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == electrolyte[2]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == electrolyte[3]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[6]
            # Chemical factory
            elif ws['C' + str(i)].value == electrolyte[4]:
                ws['E' + str(i)].value = (ws2['I' + str(a[14])].value + ws2['I' + str(a[15])].value) * \
                                         values_from_Ellingsen[1]

        # Copper - Foil (Negative current collector)
        elif ws['B' + str(i)].value == '=Foreground!$B$20':

            # Copper
            if ws['C' + str(i)].value == negative_cc_Cu[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[9])].value
            # Scrap copper
            elif ws['C' + str(i)].value == negative_cc_Cu[1]:
                ws['E' + str(i)].value = 0
            # Sheet rolling, copper
            elif ws['C' + str(i)].value == negative_cc_Cu[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[9])].value + 0
            # Transport: Freight train
            elif ws['C' + str(i)].value == negative_cc_Cu[3]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == negative_cc_Cu[4]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[6]
            # Metal working factory
            elif ws['C' + str(i)].value == negative_cc_Cu[5]:
                ws['E' + str(i)].value = (ws2['I' + str(a[9])].value + 0) * values_from_Ellingsen[4]

        # Aluminum - Foil (Positive current collector)
        elif ws['B' + str(i)].value == '=Foreground!$B$21':

            # Aluminium
            if ws['C' + str(i)].value == positive_cc_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == positive_cc_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value
            # # Transport: Freight train
            elif ws['C' + str(i)].value == positive_cc_Al[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_cc_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value * values_from_Ellingsen[6]
            # Aluminium casting facility
            elif ws['C' + str(i)].value == positive_cc_Al[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[11])].value / values_from_Ellingsen[3]

        # Negative electrode paste
        elif ws['B' + str(i)].value == '=Foreground!$B$22':

            # Carboxymethyl cellulose
            if ws['C' + str(i)].value == negative_electrode_paste[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value * values_from_Ellingsen[8]
            # Acrylic acid (PAA)
            elif ws['C' + str(i)].value == negative_electrode_paste[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value * values_from_Ellingsen[8]
            # Water - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[22])].value
            # NMP - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[20])].value
            # Carbon black - Anode
            elif ws['C' + str(i)].value == negative_electrode_paste[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[3])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == negative_electrode_paste[5]:
                ws['E' + str(i)].value = (2 * ws2['I' + str(a[1])].value * values_from_Ellingsen[8] +
                                          ws2['I' + str(a[1])].value + ws2['I' + str(a[22])].value) * \
                                         values_from_Ellingsen[7]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == negative_electrode_paste[6]:
                ws['E' + str(i)].value = (2 * ws2['I' + str(a[1])].value * values_from_Ellingsen[8] +
                                          ws2['I' + str(a[1])].value + ws2['I' + str(a[22])].value) * \
                                         values_from_Ellingsen[6]

        # Positive electrode paste
        elif ws['B' + str(i)].value == '=Foreground!$B$23':

            # NMP - Cathode
            if ws['C' + str(i)].value == positive_electrode_paste[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[21])].value
            # Carbon black - Cathode
            elif ws['C' + str(i)].value == positive_electrode_paste[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[4])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == positive_electrode_paste[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[1]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == positive_electrode_paste[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[2]

        # Anode tab
        elif ws['B' + str(i)].value == '=Foreground!$B$24':

            # Copper
            if ws['C' + str(i)].value == tab_Cu[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value
            # Scrap copper
            elif ws['C' + str(i)].value == tab_Cu[1]:
                ws['E' + str(i)].value = 0
            # Sheet rolling, copper
            elif ws['C' + str(i)].value == tab_Cu[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == tab_Cu[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == tab_Cu[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[6]
            # Metal working factory
            elif ws['C' + str(i)].value == tab_Cu[5]:
                ws['E' + str(i)].value = ws2['I' + str(a[10])].value * values_from_Ellingsen[4]

        # Cathode tab
        elif ws['B' + str(i)].value == '=Foreground!$B$25':

            # Aluminium
            if ws['C' + str(i)].value == tab_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == tab_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value
            # Transport: Freight train
            elif ws['C' + str(i)].value == tab_Al[2]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == tab_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value * values_from_Ellingsen[6]
            # Aluminium casting facility
            elif ws['C' + str(i)].value == tab_Al[4]:
                ws['E' + str(i)].value = ws2['I' + str(a[12])].value / values_from_Ellingsen[3]

        # Container
        elif ws['B' + str(i)].value == '=Foreground!$B$26':

            # Aluminum - Pouch
            if ws['C' + str(i)].value == pouch_Al[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value
            # Plastic PET
            elif ws['C' + str(i)].value == pouch_Al[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[19])].value
            # Nylon 6
            elif ws['C' + str(i)].value == pouch_Al[2]:
                ws['E' + str(i)].value = 0
            # Plastic PP
            elif ws['C' + str(i)].value == pouch_Al[3]:
                ws['E' + str(i)].value = ws2['I' + str(a[18])].value
            # Packaging film
            elif ws['C' + str(i)].value == pouch_Al[4]:
                ws['E' + str(i)].value = 0
            # Injection moulding
            elif ws['C' + str(i)].value == pouch_Al[5]:
                ws['E' + str(i)].value = ws2['I' + str(a[19])].value + ws2['I' + str(a[18])].value
            # Sheet rolling, aluminium
            elif ws['C' + str(i)].value == pouch_Al[6]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value
            # Aluminium casting facility
            elif ws['C' + str(i)].value == pouch_Al[7]:
                ws['E' + str(i)].value = ws2['I' + str(a[13])].value / values_from_Ellingsen[3]
            # Plastic processing factory
            elif ws['C' + str(i)].value == pouch_Al[8]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value) / \
                                         values_from_Ellingsen[2]
            # Transport: Freight train
            elif ws['C' + str(i)].value == pouch_Al[9]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value +
                                          ws2['I' + str(a[13])].value) * values_from_Ellingsen[5]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == pouch_Al[10]:
                ws['E' + str(i)].value = (ws2['I' + str(a[18])].value + ws2['I' + str(a[19])].value +
                                          ws2['I' + str(a[13])].value) * values_from_Ellingsen[6]

        # Negative active material
        elif ws['B' + str(i)].value == '=Foreground!$B$27':

            # Battery grade graphite - Anode
            if ws['C' + str(i)].value == negative_active_material[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[1])].value
            # Silicon nano-wire - Anode
            elif ws['C' + str(i)].value == negative_active_material[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[2])].value

        # Negative binder
        elif ws['B' + str(i)].value == '=Foreground!$B$29':

            # PVDF
            if ws['C' + str(i)].value == binders[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[7])].value
            # CMC
            elif ws['C' + str(i)].value == binders[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[5])].value

        # Positive binder
        elif ws['B' + str(i)].value == '=Foreground!$B$30':
            # PVDF
            if ws['C' + str(i)].value == binders[0]:
                ws['E' + str(i)].value = ws2['I' + str(a[8])].value
            # CMC
            elif ws['C' + str(i)].value == binders[1]:
                ws['E' + str(i)].value = ws2['I' + str(a[6])].value

        # Positive active material
        elif ws['D' + str(i)].value == '=Foreground!$C$34':

            # Sodium hydroxide
            if ws['C' + str(i)].value == NCA[0]:
                ws['E' + str(i)].value = ws['R6'].value * values_from_Ellingsen[9]
            # Lithium hydroxide
            elif ws['C' + str(i)].value == NCA[1]:
                ws['E' + str(i)].value = ws['R6'].value * values_from_Ellingsen[10]
            # Nickel
            elif ws['C' + str(i)].value == NCA[2]:
                ws['E' + str(i)].value = ws['R6'].value * values_from_Ellingsen[11]
            # Aluminium hydroxide
            elif ws['C' + str(i)].value == NCA[3]:
                ws['E' + str(i)].value = ws['R6'].value * values_from_Ellingsen[12]
            # Heat in chemical industry
            elif ws['C' + str(i)].value == NCA[4]:
                ws['E' + str(i)].value = ws['R6'].value * values_from_MajeauBettez[0]
            # Transport: Freight train
            elif ws['C' + str(i)].value == NCA[5]:
                ws['E' + str(i)].value = (values_from_Ellingsen[9]+values_from_Ellingsen[10]+values_from_Ellingsen[11]
                                          + values_from_Ellingsen[12]) * values_from_Ellingsen[7] * ws['R6'].value
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == NCA[6]:
                ws['E' + str(i)].value = (values_from_Ellingsen[9]+values_from_Ellingsen[10]+values_from_Ellingsen[11]
                                          + values_from_Ellingsen[12]) * values_from_Ellingsen[6] * ws['R6'].value
            # Chemical plant
            elif ws['C' + str(i)].value == NCA[7]:
                ws['E' + str(i)].value = values_from_Ellingsen[1]

        # Cobalt sulphate
        if ws['D' + str(i)].value == '=Foreground!$C$42':

            # Chemicals, inorganic
            if ws['C' + str(i)].value == cobalt_sulphate[0]:
                ws['E' + str(i)].value = values_from_MajeauBettez[7]
            # Chemicals, organic
            elif ws['C' + str(i)].value == cobalt_sulphate[1]:
                ws['E' + str(i)].value = values_from_MajeauBettez[8]
            # Hydrogen cyanide
            elif ws['C' + str(i)].value == cobalt_sulphate[2]:
                ws['E' + str(i)].value = values_from_MajeauBettez[9]
            # Lime
            elif ws['C' + str(i)].value == cobalt_sulphate[3]:
                ws['E' + str(i)].value = values_from_MajeauBettez[10]
            # Cement
            elif ws['C' + str(i)].value == cobalt_sulphate[4]:
                ws['E' + str(i)].value = values_from_MajeauBettez[11]
            # Sand/Gravel
            elif ws['C' + str(i)].value == cobalt_sulphate[5]:
                ws['E' + str(i)].value = values_from_MajeauBettez[12]
            # Blasting
            elif ws['C' + str(i)].value == cobalt_sulphate[6]:
                ws['E' + str(i)].value = values_from_MajeauBettez[13]
            # Diesel
            elif ws['C' + str(i)].value == cobalt_sulphate[7]:
                ws['E' + str(i)].value = values_from_MajeauBettez[14]
            # Electricity
            elif ws['C' + str(i)].value == cobalt_sulphate[8]:
                ws['E' + str(i)].value = values_from_MajeauBettez[15]
            # Aluminium hydroxide
            elif ws['C' + str(i)].value == cobalt_sulphate[9]:
                ws['E' + str(i)].value = values_from_MajeauBettez[16]
            # Conveyor belt
            elif ws['C' + str(i)].value == cobalt_sulphate[10]:
                ws['E' + str(i)].value = values_from_MajeauBettez[17]
            # Mine infrastructure
            elif ws['C' + str(i)].value == cobalt_sulphate[11]:
                ws['E' + str(i)].value = values_from_MajeauBettez[18]
            # Transport: Lorry >32 ton
            elif ws['C' + str(i)].value == cobalt_sulphate[12]:
                ws['E' + str(i)].value = values_from_MajeauBettez[19]
            # Non-sulfidic overburden
            elif ws['C' + str(i)].value == cobalt_sulphate[13]:
                ws['E' + str(i)].value = values_from_MajeauBettez[20]
            # Non-sulfidic tailing
            elif ws['C' + str(i)].value == cobalt_sulphate[14]:
                ws['E' + str(i)].value = values_from_MajeauBettez[21]

# Chemistry-independent values

for i in range(5, ws.max_row + 1):

    # Battery cell
    if ws['C' + str(i)].value == bat_cell[0]:
        ws['E' + str(i)].value = values_from_Ellingsen[0]

    # Electrolyte
    if ws['C' + str(i)].value == electrolyte[4]:
        ws['E' + str(i)].value = values_from_Ellingsen[1]


''' Electricity mixes '''
ws4 = wb['Electricity_mixes']

ws4['A1'].value = 'Country'
ws4['A2'].value = 'Coal'
ws4['A3'].value = 'Oil'
ws4['A4'].value = 'Natural Gas'
ws4['A5'].value = 'Nuclear'
ws4['A6'].value = 'Hydropower'
ws4['A7'].value = 'Solar PV'
ws4['A8'].value = 'Wind'
ws4['A9'].value = 'Waste heat'
ws4['A10'].value = 'Biofuels'
ws4['A11'].value = 'Other Sources'
ws4['A13'].value = 'SUM'

ws4['B1'].value = 'Norway'
ws4['B2'].value = 0.121
ws4['B3'].value = 0.013
ws4['B4'].value = 1.732
ws4['B5'].value = 0
ws4['B6'].value = 93.436
ws4['B7'].value = 0.010
ws4['B8'].value = 4.112
ws4['B9'].value = 0.306
ws4['B10'].value = 0.027
ws4['B11'].value = 0.244
ws4['B13'].value = '=SUM(B2:B11)'

ws4['C1'].value = 'Sweden'
ws4['C2'].value = 0.972
ws4['C3'].value = 0.163
ws4['C4'].value = 0.448
ws4['C5'].value = 39.530
ws4['C6'].value = 38.671
ws4['C7'].value = 0.394
ws4['C8'].value = 11.782
ws4['C9'].value = 3.019
ws4['C10'].value = 5.021
ws4['C11'].value = 0
ws4['C13'].value = '=SUM(C2:C11)'

ws4['D1'].value = 'Denmark'
ws4['D2'].value = 11.087
ws4['D3'].value = 0.809
ws4['D4'].value = 6.326
ws4['D5'].value = 0
ws4['D6'].value = 0.055
ws4['D7'].value = 3.289
ws4['D8'].value = 55.161
ws4['D9'].value = 5.875
ws4['D10'].value = 17.399
ws4['D11'].value = 0
ws4['D13'].value = '=SUM(D2:D11)'

ws4['E1'].value = 'Germany'
ws4['E2'].value = 30.004
ws4['E3'].value = 0.822
ws4['E4'].value = 15.279
ws4['E5'].value = 12.143
ws4['E6'].value = 4.238
ws4['E7'].value = 7.686
ws4['E8'].value = 20.376
ws4['E9'].value = 2.033
ws4['E10'].value = 7.219
ws4['E11'].value = 0.200
ws4['E13'].value = '=SUM(E2:E11)'

ws4['F1'].value = 'India'
ws4['F2'].value = 71.008
ws4['F3'].value = 0.492
ws4['F4'].value = 4.490
ws4['F5'].value = 2.916
ws4['F6'].value = 10.906
ws4['F7'].value = 3.172
ws4['F8'].value = 4.142
ws4['F9'].value = 0.098
ws4['F10'].value = 2.775
ws4['F11'].value = 0
ws4['F13'].value = '=SUM(F2:F11)'

ws4['G1'].value = 'Japan'
ws4['G2'].value = 31.643
ws4['G3'].value = 4.800
ws4['G4'].value = 33.903
ws4['G5'].value = 6.379
ws4['G6'].value = 8.802
ws4['G7'].value = 7.412
ws4['G8'].value = 0.750
ws4['G9'].value = 2.373
ws4['G10'].value = 1.757
ws4['G11'].value = 2.182
ws4['G13'].value = '=SUM(G2:G11)'

ws4['H1'].value = 'USA'
ws4['H2'].value = 24.245
ws4['H3'].value = 0.824
ws4['H4'].value = 37.435
ws4['H5'].value = 19.314
ws4['H6'].value = 6.795
ws4['H7'].value = 2.133
ws4['H8'].value = 6.949
ws4['H9'].value = 0.401
ws4['H10'].value = 1.270
ws4['H11'].value = 0.635
ws4['H13'].value = '=SUM(H2:H11)'

ws4['I1'].value = 'South Korea'
ws4['I2'].value = 40.322
ws4['I3'].value = 2.519
ws4['I4'].value = 26.044
ws4['I5'].value = 25.099     # Nuclear
ws4['I6'].value = 1.072
ws4['I7'].value = 2.240
ws4['I8'].value = 0.459
ws4['I9'].value = 0.186
ws4['I10'].value = 1.571
ws4['I11'].value = 0.487
ws4['I13'].value = '=SUM(I2:I11)'

ws4['J1'].value = 'China'
ws4['J2'].value = 64.618
ws4['J3'].value = 0.143
ws4['J4'].value = 3.332
ws4['J5'].value = 4.637
ws4['J6'].value = 17.314
ws4['J7'].value = 2.976
ws4['J8'].value = 5.395
ws4['J9'].value = 0.204
ws4['J10'].value = 1.374
ws4['J11'].value = 0.006
ws4['J13'].value = '=SUM(J2:J11)'

letters_BtoJ = [
    'B',
    'C',
    'D',
    'E',
    'F',
    'G',
    'H',
    'I',
    'J'
]

for letter in letters_BtoJ:
    ws4[letter + '13'].number_format = '0.00%'
    for i in range(2, 12):
        ws4[letter + str(i)].number_format = '0.00%'
        ws4[letter + str(i)].value = ws4[letter + str(i)].value / 100

for i, source in enumerate(electricity_sources):
    ws['I' + str(i + 7)].value = source

# ws2['I6'].value = input("Choose a country: ")
electricity_mix = input('Norway\nSweden\nDenmark\nGermany\nIndia\nJapan\nUSA\nSouth Korea\nChina\nSelect one of the '
                        'countries listed above for the electricity mix: ')
if electricity_mix.upper() not in ('NORWAY', 'SWEDEN', 'DENMARK', 'GERMANY', 'INDIA', 'JAPAN', 'USA', 'SOUTH KOREA',
                                   'CHINA'):
    print('WARNING: The chosen country is not in the list of available choices. Program stopped.')
    exit()

ws['I6'] = electricity_mix.upper()

if ws['I6'].value == 'NORWAY':
    for i in range(7, 17):
        ws['J' + str(i)].value = '=VLOOKUP(I' + str(i) + ',Electricity_mixes!$A$2:$K$11,2,FALSE)'
        ws['J' + str(i)].number_format = '0.00%'
elif ws['I6'].value == 'SWEDEN':
    for i in range(7, 17):
        ws['J' + str(i)].value = '=VLOOKUP(I' + str(i) + ',Electricity_mixes!$A$2:$K$11,3,FALSE)'
        ws['J' + str(i)].number_format = '0.00%'
elif ws['I6'].value == 'DENMARK':
    for i in range(7, 17):
        ws['J' + str(i)].value = '=VLOOKUP(I' + str(i) + ',Electricity_mixes!$A$2:$K$11,4,FALSE)'
        ws['J' + str(i)].number_format = '0.00%'
elif ws['I6'].value == 'GERMANY':
    for i in range(7, 17):
        ws['J' + str(i)].value = '=VLOOKUP(I' + str(i) + ',Electricity_mixes!$A$2:$K$11,6,FALSE)'
        ws['J' + str(i)].number_format = '0.00%'
elif ws['I6'].value == 'SOUTH KOREA':
    for i in range(7, 17):
        ws['J' + str(i)].value = '=VLOOKUP(I' + str(i) + ',Electricity_mixes!$A$2:$K$11,7,FALSE)'
        ws['J' + str(i)].number_format = '0.00%'
elif ws['I6'].value == 'JAPAN':
    for i in range(7, 17):
        ws['J' + str(i)].value = '=VLOOKUP(I' + str(i) + ',Electricity_mixes!$A$2:$K$11,8,FALSE)'
        ws['J' + str(i)].number_format = '0.00%'
elif ws['I6'].value == 'CHINA':
    for i in range(7, 17):
        ws['J' + str(i)].value = '=VLOOKUP(I' + str(i) + ',Electricity_mixes!$A$2:$K$11,9,FALSE)'
        ws['J' + str(i)].number_format = '0.00%'

for i in range(5, ws.max_row + 1):
    ws['A' + str(i)] = '=VLOOKUP(C' + str(i) + ',PRO!$A$2:$B$12917,2,FALSE)'
    ws['F' + str(i)] = '=VLOOKUP(C' + str(i) + ',PRO!$A$2:$P$12917,16,FALSE)'

for i in range(5, ws.max_row + 1):
    ws['C' + str(i)].fill = PatternFill(fill_type='solid', fgColor='CCFFCC')
    ws['D' + str(i)].fill = PatternFill(fill_type='solid', fgColor='CCFFCC')
    ws['E' + str(i)].fill = PatternFill(fill_type='solid', fgColor='CCFFCC')

ws['A1'].value = 'In this sheet, you enter the coordinates of the requirements placed on the background by the '\
                 'foreground. This will be assembled as an A_bf matrix '
ws['A3'].value = 'Background Process Name'
ws['B3'].value = 'Foreground Process Name'
ws['C3'].value = '(Matrix Row  position)'
ws['D3'].value = '(Matrix column position)'
ws['F3'].value = 'Unit'
ws['A4'].value = 'Comment'
ws['B4'].value = 'Comment'
ws['C4'].value = 'BACKGROUND PROCESS ID #'
ws['D4'].value = 'FOREGROUND PROCESS ID #'
ws['E4'].value = 'VALUE'
ws['F4'].value = 'Comment'
ws['G4'].value = 'Comment'
ws['H4'].value = 'Comment'
ws['I5'].value = 'Chosen electricity mix'
ws['R5'] = 'Mass of active material'
ws['O5'] = 'Component'
ws['P5'] = 'Weight'
ws['P6'] = '=SUM(P7:P11)'
ws['L6'] = 'Aspect'
ws['M6'] = 'Value'
ws['O6'] = 'Battery cell'
ws['L7'] = 'High voltage El required per kWh of Medium volt. El at grid'
ws['M7'] = 1.01080
ws['O7'] = 'Anode'
ws['L8'] = 'PowerPlant El required for High voltage at grid'
ws['M8'] = 1.01020
ws['O8'] = 'Cathode'
ws['O9'] = 'Separator'
ws['O10'] = 'Electrolyte'
ws['L10'] = 'SUM (HV)'
ws['M10'] = '=SUM(J24,J37,J50,J63,J76,J89,J102,J115,J128,J141)'
ws['O11'] = 'Cell container'
ws['I17'].value = 'SUM'
ws['J17'].value = '=SUM(J7:J16)'
ws['J17'].number_format = '0.00%'


wbPRO = load_workbook(filename='PRO_masterprosjekt.xlsx')
wsPRO = wbPRO.active

wbSTR = load_workbook(filename='STR_masterprosjekt.xlsx')
wsSTR = wbSTR.active


# Copying the PRO sheet from an external file into the LCI
ws5 = wb['PRO']

mrPRO = wsPRO.max_row
mcPRO = wsPRO.max_column

for i in range(1, mrPRO + 1):
    for j in range(1, mcPRO + 1):
        c = wsPRO.cell(row=i, column=j)
        ws5.cell(row=i, column=j).value = c.value

# Copying the STR sheet from an external file into the LCI
ws6 = wb['STR']

mrSTR = wsSTR.max_row
mcSTR = wsSTR.max_column

for i in range(1, mrSTR + 1):
    for j in range(1, mcSTR + 1):
        c = wsSTR.cell(row=i, column=j)
        ws6.cell(row=i, column=j).value = c.value

'''F_f part two'''
ws = wb['F_f']

stressors_NMC_NCA = [
    509,    # [0] Sulfur hexafluoride/air
    964,    # [1] Heat, waste/soil
    177,    # [2] Dinitrogen monoxide/air
    387,    # [3] Ozone/air
    720,    # [4] Sodium sulfate
    599,    # [5] Cobalt_ in ground
    800,    # [6] Water_ river
    26386,  # [7] Water/water/ground-/m3
    118,    # [8] Carbon disulfide
    240,    # [9] Heat_ waste
    396,    # [10] Particulates_ < 2.5 um
    400,    # [11] Particulates_ > 10 um
    404,    # [12] Particulates_ > 2.5 um_ and < 10um
    1168,   # [13] Aluminum
    1187,   # [14] Arsenic_ ion
    1241,   # [15] Cadmium_ ion
    1247,   # [16] Calcium_ ion
    1282,   # [17] Chromium_ ion
    1288,   # [18] Cobalt
    1235,   # [19] COD_ Chemical Oxygen Demand
    1297,   # [20] Copper_ ion
    1301,   # [21] Cyanide
    1373,   # [22] Iron_ ion
    1383,   # [23] Lead
    1389,   # [24] Manganese
    1399,   # [25] Mercury
    1428,   # [26] Nickel_ ion
    24675,  # [27] Nitrogen/water/ground-/kg
    1543,   # [28] Solved solids
    1533,   # [29] Sulfate
    1604    # [30] Zinc_ ion
]

stressors_LFP = [
    1389,
    1373,
    8257,
    26955
]

stressor_values_MajeauBettez = [
    0.50189633525553600000000,
    0.01444853086341690000000,
    0.08364938920925600000000,
    0.00418246946046280000000,
    6.42579398925648000000000,
    0.00074904225791924700000,
    0.00771855727703589000000,
    0.00672997358638105000000,
    0.00000722426543170847000,
    0.00000024676569816730500,
    0.00000002646362494983730,
    0.05741389895726210000000,
    0.00000004600716406509080,
    0.00000006539861338178200,
    0.00087451634173313100000,
    0.00000066539286870999100,
    0.00015665249251915200000,
    0.00002433436776996540000,
    0.00000023421828978591700,
    0.00000206842126044706000,
    0.00000000315586332016739,
    0.00000203800330073460000,
    0.00190872697195666000000,
    0.00043345592590250800000,
    0.19733651363456300000000,
    0.00000638777153961591000
]

stressor_values_LFP_Ellingsen = [
    9.6E-02,
    1.9E-02,
    3.2E-02,
    6.4E-01
]

j = 5
if battery_chemistry.upper() in ('NMC811', 'NMC622', 'NMC111', 'NMC532LMO', 'NMC532', 'NMC333SI', 'NCA'):
    for i in range(5, wb['A_bf'].max_row + 1):
        if wb['A_bf']['C' + str(i)].value == sulfur:
            ws['A' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$E$25951,4,FALSE)'
            ws['B' + str(j)] = wb['A_bf']['B' + str(i)].value
            ws['C' + str(j)] = stressors_NMC_NCA[0]
            ws['D' + str(j)] = wb['A_bf']['D' + str(i)].value
            ws['E' + str(j)] = '=SUM(A_bf!E' + str(i-12) + ':E' + str(i-4) + ')*0.000000075425/(1.0108*1.0102)'
            ws['F' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$M$25951,12,FALSE)'
            j += 1
            ws['A' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$E$25951,4,FALSE)'
            ws['B' + str(j)] = wb['A_bf']['B' + str(i)].value
            ws['C' + str(j)] = stressors_NMC_NCA[1]
            ws['D' + str(j)] = wb['A_bf']['D' + str(i)].value
            ws['E' + str(j)] = '=SUM(A_bf!E' + str(i - 12) + ':E' + str(i - 4) + ')*0.00178+SUM(A_bf!E' \
                               + str(i - 12) + ':E' + str(i - 4) + ')*0.0201/(1.0108*1.0102)'
            ws['F' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$M$25951,12,FALSE)'
            j += 1
            ws['A' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$E$25951,4,FALSE)'
            ws['B' + str(j)] = wb['A_bf']['B' + str(i)].value
            ws['C' + str(j)] = stressors_NMC_NCA[2]
            ws['D' + str(j)] = wb['A_bf']['D' + str(i)].value
            ws['E' + str(j)] = '=SUM(A_bf!E' + str(i - 12) + ':E' + str(i - 4) + ')*0.000005'
            ws['F' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$M$25951,12,FALSE)'
            j += 1
            ws['A' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$E$25951,4,FALSE)'
            ws['B' + str(j)] = wb['A_bf']['B' + str(i)].value
            ws['C' + str(j)] = stressors_NMC_NCA[3]
            ws['D' + str(j)] = wb['A_bf']['D' + str(i)].value
            ws['E' + str(j)] = '=SUM(A_bf!E' + str(i - 12) + ':E' + str(i - 4) + ')*0.0000045'
            ws['F' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$M$25951,12,FALSE)'
            j += 1
        if wb['A_bf']['C' + str(i)].value == 6293:
            ws['A' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$E$25951,4,FALSE)'
            ws['B' + str(j)] = wb['A_bf']['B' + str(i)].value
            ws['C' + str(j)] = stressors_NMC_NCA[4]
            ws['D' + str(j)] = wb['A_bf']['D' + str(i)].value
            ws['E' + str(j)] = 1.55683059126812
            ws['F' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$M$25951,12,FALSE)'
            j += 1
        if wb['A_bf']['C' + str(i)].value == 469:
            for k in range(5, len(stressors_NMC_NCA)):
                ws['A' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$E$25951,4,FALSE)'
                ws['B' + str(j)] = wb['A_bf']['B' + str(i)].value
                ws['C' + str(j)] = stressors_NMC_NCA[k]
                ws['D' + str(j)] = wb['A_bf']['D' + str(i)].value
                ws['E' + str(j)] = stressor_values_MajeauBettez[k-5]
                ws['F' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$M$25951,12,FALSE)'
                j += 1

if battery_chemistry.upper() == 'LFP':
    for i in range(5, len(stressors_LFP)+5):
        ws['A' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$E$25951,4,FALSE)'
        ws['B' + str(j)] = 'LFP G'
        ws['C' + str(j)] = stressors_LFP[i-5]
        ws['D' + str(j)] = wb['Foreground']['C33'].value
        ws['E' + str(j)] = stressor_values_LFP_Ellingsen[i-5]
        ws['F' + str(j)] = '=VLOOKUP(C' + str(j) + ',STR!$B$2:$M$25951,12,FALSE)'
        j += 1

'''y_gen'''
ws = wb['y_gen']
ws['A1'] = 'This spread sheet contains coordinates to build a final demand vector.'
ws['A2'] = 'Only use this if you do NOT upload a foreground.  Otherwise, use the “y_f” field in the sheet A_bf.'
ws['A4'] = 'comment'
ws['B4'] = 'comment'
ws['C4'] = 'row index of process'
ws['D4'] = 'Final demand'
ws['E4'] = 'comments'
ws['F4'] = 'comments'
ws['G4'] = 'comments'

wb.save(filename=battery_chemistry.upper() + '_' + electricity_mix + '.xlsx')
